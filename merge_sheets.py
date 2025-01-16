import openpyxl
from openpyxl import Workbook
from tqdm import tqdm
import re

def merge_excel_sheets(input_file, output_file):
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)

    # Sheets to exclude
    excluded_sheets = {'Overview', 'Content', 'Characteristics & demographics'}

    # Create a new workbook for the merged data
    new_wb = Workbook()
    merged_sheet = new_wb.active
    merged_sheet.title = "Merged Data"

    # Initialize a row counter for the merged sheet
    current_row = 1

    # Loop through the sheets in the input workbook
    for sheet_name in tqdm(wb.sheetnames, desc="Processing sheets"):
        if sheet_name in excluded_sheets:
            continue

        sheet = wb[sheet_name]

        # Copy rows from the current sheet to the merged sheet
        for row in tqdm(sheet.iter_rows(values_only=True), desc=f"Processing rows in {sheet_name}", leave=False):
            merged_sheet.append(row)

        # Add a blank row between sheets (optional for better readability)
        current_row += sheet.max_row
        merged_sheet.append([])

    # # Append question to each option in the table
    # question = None
    # for row_idx, row in enumerate(tqdm(merged_sheet.iter_rows(values_only=True), desc="Appending question to options"), start=1):
    #     if row and isinstance(row[0], str):
    #         match = re.match(r"^(.*?\?).*", row[0].strip())
    #         if match:
    #             question = match.group(1).strip()  # Capture the question up to the question mark
    #     elif question and row and isinstance(row[0], str):
    #         # Append the question to the option in the first column
    #         row_value = f"{question} {row[0]}"
    #         merged_sheet.cell(row=row_idx, column=1, value=row_value)

    # Remove 6 rows under each table based on identifying 'Survey Name' in the first column
    rows_to_delete = []
    for row_idx, row in enumerate(tqdm(merged_sheet.iter_rows(values_only=True), desc="Identifying rows to delete"), start=1):
        if row and isinstance(row[0], str) and row[0].strip().lower().startswith("survey name"):
            rows_to_delete.extend(range(row_idx, row_idx + 6))

    # Remove rows from the merged sheet
    for row_idx in tqdm(sorted(set(rows_to_delete), reverse=True), desc="Deleting rows"):
        if row_idx <= merged_sheet.max_row:
            merged_sheet.delete_rows(row_idx)

    # Remove remaining empty rows
    empty_rows = []
    for row_idx, row in enumerate(tqdm(merged_sheet.iter_rows(values_only=True), desc="Identifying empty rows"), start=1):
        if all(cell is None for cell in row):
            empty_rows.append(row_idx)

    for row_idx in tqdm(sorted(empty_rows, reverse=True), desc="Deleting empty rows"):
        if row_idx <= merged_sheet.max_row:
            merged_sheet.delete_rows(row_idx)

    # Save the merged workbook
    new_wb.save(output_file)

# Example usage
input_file = "statista_data/book-market-in-italy/advanced_reports/Food and Nutrition in Italy 2021 adv.xlsx"  # Replace with the path to your input Excel file
output_file = "merged_output.xlsx"  # Replace with the desired output file path
merge_excel_sheets(input_file, output_file)
