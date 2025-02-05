import os
import pandas as pd
import logging
import warnings
import openpyxl
from openpyxl import Workbook
import re
from datetime import datetime


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("scraper.log", mode="w", encoding="utf-8"),
    ],
)
log = logging.getLogger()


def format_relative_path(base_dir, full_path):
    """Format and return the relative path from the base directory."""
    try:
        return os.path.relpath(full_path, base_dir)
    except ValueError:
        return full_path


def process_files(selected_files, base_dir, process_function, skip_adv=False):
    """
    Generic function to process selected files using a given transformation function.
    Optionally skips files containing "adv" in their name if skip_adv is True.
    Provides detailed logs for every step of the process.
    """
    transformed_files = []
    skipped_files = []
    errors = []

    logging.info("Processing selected files...")
    logging.info(f"Base directory: {base_dir}")
    logging.info(f"Total files selected: {len(selected_files)}")

    for relative_path in selected_files:
        try:
            input_path = os.path.join(base_dir, relative_path)
            logging.info(f"Starting processing for file: {relative_path}")

            # Check if the file exists
            if not os.path.exists(input_path):
                logging.warning(
                    f"❌ File not found: {relative_path}. Skipping transformation."
                )
                errors.append((relative_path, "File not found"))
                continue

            # Skip files with 'adv' in the name if skip_adv is True
            if skip_adv and "adv" in os.path.basename(input_path).lower():
                logging.info(
                    f"⚠️ Skipping file: {relative_path} (Reason: Contains 'adv' in the name)"
                )
                skipped_files.append((relative_path, "Contains 'adv'"))
                transformed_files.append(format_relative_path(base_dir, input_path))
                continue

            # Determine the output directory for transformed files
            parent_dir = os.path.dirname(input_path)
            transformed_dir = (
                parent_dir
                if parent_dir.endswith("transformed")
                else os.path.join(parent_dir, "transformed")
            )
            os.makedirs(transformed_dir, exist_ok=True)
            logging.info(f"📂 Transformed directory created: {transformed_dir}")

            output_path = os.path.join(transformed_dir, os.path.basename(input_path))
            formatted_path = format_relative_path(base_dir, output_path)

            # Process the file using the provided transformation function
            with pd.ExcelFile(input_path) as xls:
                logging.info(f"📄 Reading file: {input_path}")
                sheet_data = process_function(xls)

            # Save the processed data
            if isinstance(
                sheet_data, Workbook
            ):  # If a Workbook is returned (e.g., merging sheets)
                logging.info(f"💾 Saving workbook to: {output_path}")
                sheet_data.save(output_path)
            elif isinstance(sheet_data, dict):  # If multiple sheets are returned
                logging.info(
                    f"💾 Saving data to: {output_path} (Multiple sheets detected)"
                )
                with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                    for sheet_name, data in sheet_data.items():
                        logging.info(f"    Writing sheet: {sheet_name}")
                        data.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                logging.error(f"❌ Unexpected file type returned for: {relative_path}")
                errors.append((relative_path, "Bad file type"))
                continue

            logging.info(f"✅ File processed successfully: {formatted_path}")
            transformed_files.append(formatted_path)

        except Exception as e:
            logging.error(f"❌ Error processing file: {relative_path} (Error: {e})")
            errors.append((relative_path, str(e)))

    # Log the transformation summary
    logging.info("\n=========== Transformation Summary ===========")
    logging.info(f"Total files selected: {len(selected_files)}")
    logging.info(f"Files successfully processed: {len(transformed_files)}")
    if transformed_files:
        logging.info("Transformed files:")
        for path in transformed_files:
            logging.info(f"  - {path}")
    if skipped_files:
        logging.info(f"Files skipped: {len(skipped_files)}")
        for path, reason in skipped_files:
            logging.info(f"  - {path} (Reason: {reason})")
    if errors:
        logging.error(f"Files with errors: {len(errors)}")
        for path, error in errors:
            logging.error(f"  - {path}: {error}")
    logging.info("=============================================")

    return transformed_files


def tr1_remove_sheets(selected_files, base_dir="statista_data"):
    def process_function(xls):
        sheets = xls.sheet_names
        processed_sheets = {}

        for sheet in sheets:
            if sheet not in ["Overview", "Content", "Lists"]:
                data = pd.read_excel(xls, sheet_name=sheet, header=None)

                # Clean all sheets by removing superscripts and converting to numbers if possible
                data = data.applymap(
                    lambda x: (
                        re.sub(r"[\u00B9\u00B2\u00B3\u2070\u2074-\u2079]", "", str(x))
                        if isinstance(x, str)
                        else x
                    )
                )
                data = data.applymap(
                    lambda x: (
                        pd.to_numeric(x, errors="ignore")
                        if isinstance(x, str) and x.isdigit()
                        else x
                    )
                )

                processed_sheets[sheet] = data

        return processed_sheets

    return process_files(selected_files, base_dir, process_function)


def tr2_remove_header_and_empty_column(selected_files, base_dir="statista_data"):
    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)

            # Remove metadata rows (rows without numeric data)
            valid_data_start_index = df[
                df.apply(
                    lambda col: col.map(lambda x: isinstance(x, (int, float))), axis=0
                ).any(axis=1)
            ].index.min()

            df_cleaned = df.iloc[valid_data_start_index:].reset_index(drop=True)

            # Use the first valid row as headers
            df_cleaned.columns = df_cleaned.iloc[0]
            df_cleaned = df_cleaned[1:].reset_index(drop=True)

            # Drop entirely empty columns
            df_cleaned = df_cleaned.dropna(axis=1, how="all")

            # Replace NaN or placeholder headers
            df_cleaned.columns = [
                f"Column_{i}" if pd.isna(col) else col
                for i, col in enumerate(df_cleaned.columns)
            ]

            sheet_data[sheet] = df_cleaned
        return sheet_data

    return process_files(selected_files, base_dir, process_function, skip_adv=True)


def tr3_remove_metadata(selected_files, base_dir="statista_data"):
    keywords = [
        "Survey Name:",
        "Base n =",
        "Question Type:",
        "Sample Size n =",
        "Population:",
        "¹Low base:",
        "Survey period",
        "Survey name:",
        "Sample size n = ",
    ]

    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            df_cleaned = df[
                ~df.apply(
                    lambda row: row.astype(str)
                    .str.contains("|".join(keywords), na=False)
                    .any(),
                    axis=1,
                )
            ]
            sheet_data[sheet] = df_cleaned
        return sheet_data

    return process_files(selected_files, base_dir, process_function)


def tr4_reduce_empty_lines(selected_files, base_dir="statista_data"):
    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)

            # Replace consecutive empty rows with a single empty row
            df_cleaned = df.loc[
                (df.shift(1).isnull().all(axis=1) & df.isnull().all(axis=1)) == False
            ]

            sheet_data[sheet] = df_cleaned
        return sheet_data

    return process_files(selected_files, base_dir, process_function)


def tr5_removing_total_percentages_income_demography(
    selected_files, base_dir="statista_data"
):
    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)

            # Remove columns where any cell contains specified strings
            column_keywords = [
                "Grand Total",
                "in %",
                "Up to € 26 400",
                "€26 400 up to €50 400",
                "€50 400 up to €117 600",
                "€117 600 and more",
                "Prefer not to say" "Less than 18 000 €",
                "18 000 € up to less than 21 600 €",
                "21 600 € up to less than 26 400 €",
                "26 400 € up to less than 36 000 €",
                "36 000 € up to less than 50 400 €",
                "50 400 € up to less than 69 600 €",
                "69 600 € up to less than 91 200 €",
                "91 200 € up to less than 117 600 €",
                "117 600 € to less than 148 800 €",
                "More than 148 800 €",
                "Would not like to answer",
                "Up to 26 400€",
                "26 400€ up to 50 400€",
                "50 400€ up to 117 600€",
                "117 600€ and more",
            ]
            columns_to_keep = df.apply(
                lambda col: ~col.astype(str)
                .str.contains("|".join(map(re.escape, column_keywords)), na=False)
                .any()
            )
            df_cleaned = df.loc[:, columns_to_keep]

            # Remove rows where any cell contains specified strings
            row_keywords = ["Gender", "Age (basic)"]
            rows_to_keep = ~df_cleaned.astype(str).apply(
                lambda row: row.str.contains(
                    "|".join(map(re.escape, row_keywords)), na=False
                ).any(),
                axis=1,
            )
            df_cleaned = df_cleaned.loc[rows_to_keep]

            sheet_data[sheet] = df_cleaned
        return sheet_data

    return process_files(selected_files, base_dir, process_function)


import pandas as pd
import re


def tr6_append_questions(selected_files, base_dir="statista_data"):
    def process_function(xls):
        sheet_data = {}

        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            df.reset_index(drop=True, inplace=True)

            question = None
            rows_to_drop = []

            # Process the DataFrame for the current sheet
            for idx, row in df.iterrows():
                first_col = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""

                # Identify and extract question or statement rows
                match = re.match(
                    r"^(.*?\?|.*\(multi-pick\)|.*\(single-pick\)|Recode based on .*|Agreement with the statement:.*|Thinking about .*|Usage frequency .*|Usage intensity .*)",
                    first_col,
                )
                if match:
                    question = match.group(1).strip()

                    # Extract the part after "/" if present
                    if "/" in first_col:
                        parts = [p.strip() for p in first_col.split("/", 1)]
                        question = "_".join(parts)  # Join both parts with underscores

                    # Remove special characters and replace spaces with underscores
                    question = re.sub(r"[,:'\".]", "", question).replace(" ", "_") + "_"
                    rows_to_drop.append(
                        idx
                    )  # Collect the index of the question or statement row
                    print(f"Question: {question}")
                    continue

                # Append question or statement to the first column of option rows
                if question and first_col:
                    first_col = re.sub(r"[,:'\".]", "", first_col).replace(
                        " ", "_"
                    )  # Process the answer as well
                    df.iloc[idx, 0] = (
                        f"{question}{first_col}"  # Ensure no space before answer
                    )

            # Identify and remove rows with leftover questions
            for idx, row in df.iterrows():
                if any(
                    val in row.values for val in ["Male", "Female", "male", "female"]
                ):
                    # Check the row above for emptiness
                    if idx > 0 and not any(df.iloc[idx - 1, 1:].notna()):
                        rows_to_drop.append(idx - 1)

            # Drop identified rows
            df.drop(index=rows_to_drop, inplace=True)
            df.reset_index(drop=True, inplace=True)

            # Add the processed DataFrame to the output dictionary
            sheet_data[sheet] = df

        return sheet_data

    return process_files(selected_files, base_dir, process_function)


def tr7_merging_sheets(selected_files, base_dir="statista_data"):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    def process_function(xls):
        wb = openpyxl.load_workbook(xls)
        new_wb = Workbook()
        merged_sheet = new_wb.active
        merged_sheet.title = "Merged Data"
        current_row = 1

        for sheet_name in wb.sheetnames:

            sheet = wb[sheet_name]

            # Copy rows from the current sheet to the merged sheet
            for row in sheet.iter_rows(values_only=True):
                merged_sheet.append(row)

            # Add a blank row between sheets (optional for better readability)
            current_row += sheet.max_row
            merged_sheet.append([])

        # Remove the first row (if deemed unnecessary)
        merged_sheet.delete_rows(1)

        # Convert merged sheet to a Pandas DataFrame for easier processing
        import pandas as pd

        data = [[cell.value for cell in row] for row in merged_sheet.iter_rows()]
        df = pd.DataFrame(data)

        # Identify gender rows and remove the row above if it is not empty
        gender_rows = df[df.iloc[:, 1].isin(["Female", "Male", "male", "female"])].index
        rows_to_remove = [
            idx - 1
            for idx in gender_rows
            if idx - 1 >= 0 and not df.iloc[idx - 1].isnull().all()
        ]
        df = df.drop(index=rows_to_remove).reset_index(drop=True)

        # Write the cleaned DataFrame back to the merged sheet
        merged_sheet.delete_rows(1, merged_sheet.max_row)
        for row in dataframe_to_rows(df, index=False, header=False):
            merged_sheet.append(row)

        return new_wb

    return process_files(selected_files, base_dir, process_function)


def tr8_join_tables(selected_files, base_dir="statista_data"):
    """
    Remove all empty rows from the tables and ensure the first row with demography data is present only once.
    Additionally, if any row contains both "Male" and "Female", ensure it appears only once and remove duplicates.
    """

    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)

            # Remove completely empty rows
            df_cleaned = df.dropna(how="all").reset_index(drop=True)

            # Ensure demography row (e.g., headers or first row) appears only once
            if not df_cleaned.empty:
                first_row = df_cleaned.iloc[0]
                duplicates = df_cleaned.apply(lambda row: row.equals(first_row), axis=1)
                df_cleaned = df_cleaned.loc[~duplicates.shift(-1, fill_value=False)]

            # Identify and ensure rows with both "Male" and "Female" appear only once
            gender_rows = df_cleaned[
                df_cleaned.apply(
                    lambda row: any(
                        gender in row.values
                        for gender in ["Male", "Female", "male", "female"]
                    ),
                    axis=1,
                )
            ]
            if not gender_rows.empty:
                # Keep only the first occurrence of such rows
                unique_gender_rows = gender_rows.drop_duplicates()

                # Remove all duplicates of gender rows from the main DataFrame
                df_cleaned = df_cleaned[
                    ~df_cleaned.apply(
                        lambda row: any(
                            gender in row.values
                            for gender in ["Male", "Female", "male", "female"]
                        ),
                        axis=1,
                    )
                ]

                # Add back the unique gender rows
                df_cleaned = pd.concat(
                    [df_cleaned, unique_gender_rows], ignore_index=True
                )

            sheet_data[sheet] = df_cleaned

            # replace "Unnamed: 0" cell with "Topic"
            if "Unnamed: 0" in df_cleaned.columns:
                df_cleaned.rename(columns={"Unnamed: 0": "Topic"}, inplace=True)

        return sheet_data

    return process_files(selected_files, base_dir, process_function)


def tr9_transpose_table(selected_files, base_dir="statista_data"):
    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)

            # Transpose the data
            df_transposed = df.T.reset_index()

            # Rename columns to make transposed data consistent
            df_transposed.columns = [
                f"Column_{i}" if i > 0 else "Index"
                for i in range(df_transposed.shape[1])
            ]

            # Keep only columns with 4 or more non-empty values
            if "adv" in os.path.basename(xls).lower():
                df_filtered = df_transposed.loc[
                    :, df_transposed.notna().sum(axis=0) >= 4
                ]
            else:
                df_filtered = df_transposed

            # Update the processed data with filtered columns
            sheet_data[sheet] = df_filtered

        return sheet_data

    return process_files(selected_files, base_dir, process_function)


def tr10_map_age(selected_files, base_dir="statista_data"):
    """
    Dynamically map generational categories to specific age ranges in the format "x - y years".
    """
    from datetime import datetime

    current_year = datetime.now().year

    def map_generation_to_age(value):
        # Extended generation mapping including both formats
        generation_mapping = {
            "Gen Z (1995-2012)": (1995, 2012),
            "Millennials (1980-1994)": (1980, 1994),
            "Gen X (1965-1979)": (1965, 1979),
            "Baby Boomers (1946-1964)": (1946, 1964),
            "iGen / Gen Z (1995-2012)": (1995, 2012),
            "iGen / Generation Z (1995-2012)": (1995, 2012),
            "Millennials / Generation Y (1980-1994)": (1980, 1994),
            "Generation X (Baby Bust) (1965-1979)": (1965, 1979),
            "Baby Boomer (1946-1964)": (1946, 1964),
            "Traditionals (1922-1945)": (1922, 1945),
        }

        if value in generation_mapping:
            start_year, end_year = generation_mapping[value]
            start_age = current_year - end_year
            end_age = current_year - start_year
            return f"{start_age}-{end_age} years"

        return value  # Default to the original value if no match

    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)

            # Apply mapping to all columns where generation categories might appear
            df = df.applymap(map_generation_to_age)

            sheet_data[sheet] = df

        return sheet_data

    return process_files(selected_files, base_dir, process_function)


def tr11_filter_advanced_files(selected_files, base_dir="statista_data"):
    """
    Process files containing 'adv' in the filename to only include rows where the first column contains:
    - "Female"
    - "Male"
    - Values containing the substring "years".
    Skip filtering for the first row.
    Also, removes unwanted superscript or exponent characters from numeric values (e.g., '15¹' -> '15').
    """

    def is_valid_row(value):
        # Check if the value is "Female", "Male", or contains "years"
        return isinstance(value, str) and (
            value.lower() in ["female", "male"] or "years" in value
        )

    def clean_exponent(value):
        """
        Remove superscripts or unwanted characters like exponents (e.g., '15¹' -> '15').
        """
        if isinstance(value, str):
            # Remove any superscript numbers or other unwanted characters
            return re.sub(r"[¹²³⁴⁵⁶⁷⁸⁹⁰]", "", value)
        return value

    def process_function(xls):
        sheet_data = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, header=1, sheet_name=sheet)

            # Remove superscripts or exponents from all cells
            df = df.applymap(clean_exponent)

            # Skip the first row and apply the filter to subsequent rows
            first_row = df.iloc[0:1]
            filtered_rows = df.iloc[1:][df.iloc[1:, 0].apply(is_valid_row)]

            # Combine the first row with the filtered rows
            df_filtered = pd.concat([first_row, filtered_rows], ignore_index=True)

            sheet_data[sheet] = df_filtered

        return sheet_data

    # Only process files with "adv" in their filename
    advanced_files = [
        file for file in selected_files if "adv" in os.path.basename(file).lower()
    ]
    return process_files(advanced_files, base_dir, process_function)


def tr12_transform_to_probability(selected_files, base_dir="statista_data"):
    """
    Convert occurrences in columns to probabilities relative to their corresponding "_ Base" column,
    while preserving the first column and dropping all base columns after processing.
    Additionally, transform data to include gender-age combinations while maintaining probability integrity.
    Outputs the result as a CSV file or a pandas DataFrame.
    """

    def process_function(xls, file_path):
        sheet_data = {}

        for sheet in xls.sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name=sheet)
                df = df.loc[
                    :,
                    ~df.columns.astype(str)
                    .str.replace(".", "", regex=False)
                    .str.isnumeric(),
                ]

                # Identify "_ Base" columns
                base_columns_corrected = [col for col in df.columns if "_ Base" in col]

                # Separate the first column (preserve as-is)
                first_column = df.iloc[:, 0]  # Keep the first column intact
                remaining_data = df.iloc[:, 1:]  # Remaining columns for processing

                # Convert all relevant columns to numeric where possible (ignoring non-numeric data)
                data_numeric = remaining_data.apply(pd.to_numeric, errors="coerce")
                logging.debug(
                    f"Data converted to numeric for sheet '{sheet}':\n{data_numeric.head()}"
                )

                # Process each "_ Base" column group
                for i, base_col in enumerate(base_columns_corrected):
                    base_values_corrected = data_numeric[base_col]

                    # Determine the group of columns to process
                    if i < len(base_columns_corrected) - 1:
                        group_columns_corrected = data_numeric.columns[
                            data_numeric.columns.get_loc(base_col)
                            + 1 : data_numeric.columns.get_loc(
                                base_columns_corrected[i + 1]
                            )
                        ]
                    else:
                        group_columns_corrected = data_numeric.columns[
                            data_numeric.columns.get_loc(base_col) + 1 :
                        ]

                    # Convert occurrences to probabilities
                    for col in group_columns_corrected:
                        try:
                            data_numeric[col] = (
                                data_numeric[col] / base_values_corrected
                            )
                        except ZeroDivisionError:
                            logging.warning(
                                f"ZeroDivisionError in column '{col}' for base column '{base_col}'. Setting NaN."
                            )
                            data_numeric[col] = None

                # Drop all "_ Base" columns after processing
                data_numeric = data_numeric.drop(columns=base_columns_corrected)

                # Concatenate the preserved first column back with the transformed data
                final_df = pd.concat([first_column, data_numeric], axis=1)

                # Perform gender-age transformation
                gender_proportions = (
                    final_df.iloc[:2, 1:].astype(float).reset_index(drop=True)
                )
                gender_proportions.index = ["female", "male"]

                age_probabilities = final_df.iloc[2:].reset_index(drop=True)
                age_probabilities.rename(columns={"Topic": "age"}, inplace=True)

                transformed_data = []
                for _, row in age_probabilities.iterrows():
                    age_group = row["age"]
                    new_row_male = {"gender": "male", "age": age_group}
                    new_row_female = {"gender": "female", "age": age_group}

                    for column in gender_proportions.columns:
                        prob = row[column]  # Original probability for age group
                        female_ratio = gender_proportions.loc["female", column]
                        male_ratio = gender_proportions.loc["male", column]

                        female_value = prob * female_ratio
                        male_value = prob * male_ratio

                        total = female_value + male_value
                        if total > 0:
                            female_value = (female_value / total) * prob
                            male_value = (male_value / total) * prob

                        new_row_female[column] = female_value
                        new_row_male[column] = male_value

                    transformed_data.append(new_row_female)
                    transformed_data.append(new_row_male)

                final_transformed_df = pd.DataFrame(transformed_data)

                # Store the transformed DataFrame for this sheet
                sheet_data[sheet] = final_transformed_df
                logging.info(f"Finished processing sheet '{sheet}'")

            except Exception as e:
                logging.error(f"Error processing sheet '{sheet}': {e}")
                raise e

        # Combine all sheets into one DataFrame (optional if you want multi-sheet handling)
        combined_df = pd.concat(sheet_data.values(), ignore_index=True)
        # remove "years" from cells in age column
        combined_df["age"] = combined_df["age"].str.replace(" years", "")

        # Save to CSV
        csv_output_path = file_path.replace(".xlsx", "_transformed.csv")
        combined_df.to_csv(csv_output_path, index=False)
        logging.info(f"Transformed data saved as CSV: {csv_output_path}")

        # Remove the original `.xlsx` file
        if os.path.exists(file_path):
            os.remove(file_path)
            logging.info(f"Original file removed: {file_path}")

        return combined_df

    def process_files(files, base_dir, process_function):
        processed_dataframes = {}

        for file in files:
            file_path = os.path.join(base_dir, file)
            logging.info(f"Processing file: {file_path}")

            if os.path.exists(file_path):
                try:
                    with pd.ExcelFile(file_path) as xls:
                        processed_dataframes[file] = process_function(xls, file_path)
                except Exception as e:
                    logging.error(f"Error processing file '{file_path}': {e}")
            else:
                logging.warning(f"File not found: {file_path}")

        return processed_dataframes

    # Process only files with "adv" in their filenames
    advanced_files = [
        file for file in selected_files if "adv" in os.path.basename(file).lower()
    ]
    logging.info(f"Advanced files to process: {advanced_files}")
    return process_files(advanced_files, base_dir, process_function)


def pipeline_transform(selected_files, base_dir="statista_data"):
    logging.info("Starting the transformation pipeline...")

    # Step 1: Remove Overview
    logging.info("Step 1: Removing Overview sheets...")
    transformed_step1 = tr1_remove_sheets(selected_files, base_dir)
    if not transformed_step1:
        logging.warning("No files to process after Step 1. Exiting pipeline.")
        return

    # Step 2: Remove header and empty columns
    logging.info("Step 2: Removing header rows and empty columns...")
    transformed_step2 = tr2_remove_header_and_empty_column(transformed_step1, base_dir)
    if not transformed_step2:
        logging.warning("No files to process after Step 2. Exiting pipeline.")
        return

    # Step 3: Remove metadata rows
    logging.info("Step 3: Removing metadata rows...")
    transformed_step3 = tr3_remove_metadata(transformed_step2, base_dir)

    # Step 4: Reduce empty lines
    logging.info("Step 4: Reducing empty lines...")
    transformed_step4 = tr4_reduce_empty_lines(transformed_step3, base_dir)

    # Step 5: Remove total percentages columns
    logging.info("Step 5: Removing columns with 'Grand Total' and 'in %'...")
    transformed_step5 = tr5_removing_total_percentages_income_demography(
        transformed_step4, base_dir
    )

    # Step 6: Append questions to options
    logging.info("Step 6: Appending questions to options...")
    transformed_step6 = tr6_append_questions(transformed_step5, base_dir)

    # Step 7: Merge sheets
    logging.info("Step 6: Merging sheets...")
    transformed_step7 = tr7_merging_sheets(transformed_step6, base_dir)

    # Step 8: Join tables
    logging.info("Step 8: Joining tables...")
    transformed_step8 = tr8_join_tables(transformed_step7, base_dir)

    # Step 9: Transpose tables
    logging.info("Step 9: Transposing tables...")
    transformed_step9 = tr9_transpose_table(transformed_step8, base_dir)

    # Step 10: Map age categories
    logging.info("Step 10: Mapping age categories...")
    transformed_step10 = tr10_map_age(transformed_step9, base_dir)

    # Step 11: Filter advanced files
    logging.info("Step 11: Filtering advanced files...")
    transformed_step11 = tr11_filter_advanced_files(transformed_step10, base_dir)

    # Step 12: Transform to probability
    logging.info("Step 12: Transforming to probability...")
    transformed_step12 = tr12_transform_to_probability(transformed_step11, base_dir)

    logging.info("Transformation pipeline completed.")
    return transformed_step12
